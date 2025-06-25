from odoo import models, fields, api
import base64
import io
from openpyxl import load_workbook
from datetime import datetime

class ButtonUpload(models.TransientModel):
    _name = 'button.upload'
    _description = 'Upload Excel to Update Patients'

    file = fields.Binary(string="Upload File", required=True)
    filename = fields.Char(string="File Name")

    def upload_excel_file(self):
        if not self.file:
            return

        data = base64.b64decode(self.file)
        file_stream = io.BytesIO(data)

        workbook = load_workbook(filename=file_stream, read_only=True)
        sheet = workbook.active

        Hospital = self.env['hospital.management']
        Currency = self.env['res.currency']

        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            name, dob, marital_status, gender, address, currency = row

            if isinstance(dob, datetime):
                dob = dob.date()

            
            existing_patient = Hospital.search([
                ('name', '=', name),
                ('Address', '=', address)
            ], limit=1)

            currency_id = Currency.search([('name', '=', currency)], limit=1).id

            vals = {
                'name': name,
                'Date_of_Birth': dob,
                'Marital_Status': marital_status,
                'gender': gender,
                'Address': address,
                'currency_id': currency_id,
            }

            if existing_patient:
                existing_patient.write(vals)
            else:
                Hospital.create(vals)

        return {
            'type': 'ir.actions.client',
            'tag': 'reload',
        }

    
    
    
  
